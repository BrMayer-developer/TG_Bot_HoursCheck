import os
import openpyxl
if os.path.exists("Таблица.xlsx") == False:
    wb = openpyxl.Workbook()
    list1 = wb.create_sheet("Понедельник")
    list2 = wb.create_sheet("Вторник")
    list3 = wb.create_sheet("Среда")
    list4 = wb.create_sheet("Четверг")
    list5 = wb.create_sheet("Пятница")
    for i1 in range(5):
        i1 = str(i1 + 1)
        n1 = "A" + i1
        n2 = "B" + i1
        n = n1 + ":" + n2
        list1.merge_cells(n)
        list2.merge_cells(n)
        list3.merge_cells(n)
        list4.merge_cells(n)
        list5.merge_cells(n)

    for i2 in range(5):
        i2 = str(i2 + 1)
        n1 = "C" + i2
        n2 = "D" + i2
        n = n1 + ":" + n2
        list1.merge_cells(n)
        list2.merge_cells(n)
        list3.merge_cells(n)
        list4.merge_cells(n)
        list5.merge_cells(n)

    for i3 in range(5):
        i3 = str(i3 + 1)
        n1 = "E" + i3
        n2 = "F" + i3
        n = n1 + ":" + n2
        list1.merge_cells(n)
        list2.merge_cells(n)
        list3.merge_cells(n)
        list4.merge_cells(n)
        list5.merge_cells(n)

    for i4 in range(5):
        i4 = str(i4 + 1)
        n1 = "G" + i4
        n2 = "H" + i4
        n = n1 + ":" + n2
        list1.merge_cells(n)
        list2.merge_cells(n)
        list3.merge_cells(n)
        list4.merge_cells(n)
        list5.merge_cells(n)

    indeksi = [list1,list2,list3,list4,list5]
    for i in indeksi:
        i["A2"] = "Дмитрий"
        i["A3"] = "Ольга"
        i["A4"] = "Лариса"
        i["A5"] = "Светлана"
        i["C1"] = "Время захода"
        i["E1"] = "Время выхода"
        i["G1"] = "Количество часов"
    wb.save('Таблица.xlsx')
else:
    wb = openpyxl.open("Таблица.xlsx",data_only=True)
    list = wb.active
    wb1 = openpyxl.open("Неделя.xlsx")
    list1 = wb1.get_sheet_by_name("Sheet")
    sp = ["С","E","G","I"]
    sp1 = ["C", "E"]
    for a in sp:
        for i in range(2,6):
            p = a + str(i)
            list[p] = 0