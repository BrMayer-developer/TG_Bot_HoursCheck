import telebot,time,os,datetime
from telebot import types
import openpyxl

def creatingfile():
    global wb
    global list
    global wb1
    global list1
    if os.path.exists("Таблица.xlsx") == False:
        wb = openpyxl.Workbook()
        list = wb.active
        for i1 in range(5):
            i1 = str(i1 + 1)
            n1 = "A" + i1
            n2 = "B" + i1
            n = n1 + ":" + n2
            list.merge_cells(n)

        for i2 in range(5):
            i2 = str(i2 + 1)
            n1 = "C" + i2
            n2 = "D" + i2
            n = n1 + ":" + n2
            list.merge_cells(n)

        for i3 in range(5):
            i3 = str(i3 + 1)
            n1 = "E" + i3
            n2 = "F" + i3
            n = n1 + ":" + n2
            list.merge_cells(n)

        for i4 in range(5):
            i4 = str(i4 + 1)
            n1 = "G" + i4
            n2 = "H" + i4
            n = n1 + ":" + n2
            list.merge_cells(n)

        for i5 in range(5):
            i5 = str(i5 + 1)
            n1 = "I" + i5
            n2 = "J" + i5
            n = n1 + ":" + n2
            list.merge_cells(n)

        list["A2"] = "Дмитрий"
        list["A3"] = "Ольга"
        list["A4"] = "Лариса"
        list["A5"] = "Светлана"
        list["C1"] = "Время захода"
        list["E1"] = "Время выхода"
        list["G1"] = "Количество часов"
        list["I1"] = "Количество минут"
    else:
        wb = openpyxl.open("Таблица.xlsx")
        list = wb.get_sheet_by_name("Sheet")

    if os.path.exists("Неделя.xlsx") == False:
        wb1 = openpyxl.Workbook()
        list1 = wb1.active
        for i1 in range(5):
            i1 = str(i1 + 1)
            n1 = "A" + i1
            n2 = "B" + i1
            n = n1 + ":" + n2
            list1.merge_cells(n)

        for i2 in range(5):
            i2 = str(i2 + 1)
            n1 = "C" + i2
            n2 = "D" + i2
            n = n1 + ":" + n2
            list1.merge_cells(n)

        for i3 in range(5):
            i3 = str(i3 + 1)
            n1 = "E" + i3
            n2 = "F" + i3
            n = n1 + ":" + n2
            list1.merge_cells(n)

        list1["A2"] = "Дмитрий"
        list1["A3"] = "Ольга"
        list1["A4"] = "Лариса"
        list1["A5"] = "Светлана"
        list1["C1"] = "Количество часов"
        list1["E1"] = "Количество минут"
        list1["C2"] = 0
        list1["C3"] = 0
        list1["C4"] = 0
        list1["C5"] = 0
        list1["E2"] = 0
        list1["E3"] = 0
        list1["E4"] = 0
        list1["E5"] = 0

    else:
        wb1 = openpyxl.open("Неделя.xlsx")
        list1 = wb1.get_sheet_by_name("Sheet")


#print("Введите токен:")
#botoken = input()
botoken = "1165779184:AAELE6Gs19G6Yi-ARpCn3F883gz9816onNs"
bot = telebot.TeleBot(botoken)
tconv = lambda x: time.strftime("%H:%M:%S", time.localtime(x))


@bot.message_handler(commands=['start'])
def main(message):
    creatingfile()
    global spisok
    global spisokr
    global spisokp
    global spisoku
    global spisokpm
    global spisokum
    spisok = ["пришёл", "пришла", "пришла", "пришла"]
    spisokr = ["","","",""]
    spisokp = [0,0,0,0]
    spisoku = [0,0,0,0]
    spisokpm = [0,0,0,0]
    spisokum = [0,0,0,0]
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True,row_width=4)
    button1 = types.KeyboardButton("Дмитрий")
    button2 = types.KeyboardButton("Ольга")
    button3 = types.KeyboardButton("Светлана")
    button4 = types.KeyboardButton("Лариса")
    markup.add(button1,button2,button3,button4)
    bot.send_message(message.chat.id, "Готов к работе!", reply_markup=markup)

@bot.message_handler(commands=['excel'])
def excel(message):
    wb.save('Таблица.xlsx')
    wb1.save("Неделя.xlsx")
    f1 = open("Таблица.xlsx", "rb")
    f2 = open("Неделя.xlsx", "rb")
    bot.send_document(message.chat.id, f1)
    bot.send_document(message.chat.id, f2)

@bot.message_handler(content_types=['text'])
def func(message):
    if message.text=="Дмитрий":
        bot.delete_message(message.chat.id, message.message_id)
        bot.send_message(message.chat.id, "Дмитрий - " + spisok[0] + " в " + tconv(message.date))
        if spisok[0] == "пришёл":
            timep = tconv(message.date)
            spisokp[0] = int(timep[0:2])
            spisokpm[0] = int(timep[3:5])
            list["C2"] = timep
            spisok[0] = "ушёл"
            spisokr[0] = "пришёл"
        elif spisok[0] == "ушёл":
            timeu = tconv(message.date)
            spisoku[0] = int(timeu[0:2])
            hp = spisoku[0]-spisokp[0]
            list["G2"] = hp
            spisokum[0] = int(timeu[3:5])
            mp = abs(spisokum[0] - spisokpm[0])
            list["I2"] = mp
            list["E2"] = timeu
            hp1 = int(list1["C2"].value)
            hp1 = hp1 + hp
            list1["C2"] = hp1
            mp1 = int(list1["E2"].value)
            mp1 = mp1 + mp
            list1["E2"] = mp1
            spisok[0] = "пришёл"
            spisokr[0] = "ушёл"
            if spisokr[0]=="ушёл" and spisokr[1]=="ушла" and spisokr[2]=="ушла" and spisokr[3]=="ушла":
                wb.save('Таблица.xlsx')
                f = open("Таблица.xlsx", "rb")
                bot.send_document(message.chat.id, f)
                f.close()
                os.remove("Таблица.xlsx")
                creatingfile()
                if datetime.datetime.today().weekday()==2:
                    wb1.save('Неделя.xlsx')
                    f = open("Неделя.xlsx", "rb")
                    bot.send_document(message.chat.id, f)
                    f.close()
                    os.remove("Неделя.xlsx")
                    creatingfile()
    elif message.text == "Ольга":
        bot.delete_message(message.chat.id, message.message_id)
        bot.send_message(message.chat.id, "Ольга - " + spisok[1] + " в " + tconv(message.date))
        if spisok[1] == "пришла":
            timep = tconv(message.date)
            spisokp[1] = int(timep[0:2])
            spisokpm[1] = int(timep[3:5])
            list["C3"] = timep
            spisok[1] = "ушла"
            spisokr[1] = "пришла"
        elif spisok[1] == "ушла":
            timeu = tconv(message.date)
            spisoku[1] = int(timeu[0:2])
            hp = spisoku[1] - spisokp[1]
            list["G3"] = hp
            spisokum[1] = int(timeu[3:5])
            mp = abs(spisokum[1] - spisokpm[1])
            list["I3"] = mp
            list["E3"] = timeu
            hp1 = int(list1["C3"].value)
            hp1 = hp1 + hp
            list1["C3"] = hp1
            mp1 = int(list1["E3"].value)
            mp1 = mp1 + mp
            list1["E3"] = mp1
            spisok[1] = "пришла"
            spisokr[1] = "ушла"
            if spisokr[0]=="ушёл" and spisokr[1]=="ушла" and spisokr[2]=="ушла" and spisokr[3]=="ушла":
                wb.save('Таблица.xlsx')
                f = open("Таблица.xlsx", "rb")
                bot.send_document(message.chat.id, f)
                f.close()
                os.remove("Таблица.xlsx")
                creatingfile()
                if datetime.datetime.today().weekday()==4:
                    wb1.save('Неделя.xlsx')
                    f = open("Неделя.xlsx", "rb")
                    bot.send_document(message.chat.id, f)
                    f.close()
                    os.remove("Неделя.xlsx")
                    creatingfile()

    elif message.text=="Светлана":
        bot.delete_message(message.chat.id, message.message_id)
        bot.send_message(message.chat.id, "Светлана - " + spisok[2] + " в " + tconv(message.date))
        if spisok[2] == "пришла":
            timep = tconv(message.date)
            spisokp[2] = int(timep[0:2])
            spisokpm[2] = int(timep[3:5])
            list["C4"] = timep
            spisok[2] = "ушла"
            spisokr[2] = "пришла"
        elif spisok[2] == "ушла":
            timeu = tconv(message.date)
            spisoku[2] = int(timeu[0:2])
            hp = spisoku[2] - spisokp[2]
            list["G4"] = hp
            spisokum[2] = int(timeu[3:5])
            mp = abs(spisokum[2] - spisokpm[2])
            list["I4"] = mp
            list["E4"] = timeu
            hp1 = int(list1["C4"].value)
            hp1 = hp1 + hp
            list1["C4"] = hp1
            mp1 = int(list1["E4"].value)
            mp1 = mp1 + mp
            list1["E4"] = mp1
            spisok[2] = "пришла"
            spisokr[2] = "ушла"
            if spisokr[0]=="ушёл" and spisokr[1]=="ушла" and spisokr[2]=="ушла" and spisokr[3]=="ушла":
                wb.save('Таблица.xlsx')
                f = open("Таблица.xlsx", "rb")
                bot.send_document(message.chat.id, f)
                f.close()
                os.remove("Таблица.xlsx")
                creatingfile()
                if datetime.datetime.today().weekday()==4:
                    wb1.save('Неделя.xlsx')
                    f = open("Неделя.xlsx", "rb")
                    bot.send_document(message.chat.id, f)
                    f.close()
                    os.remove("Неделя.xlsx")
                    creatingfile()
    elif message.text=="Лариса":
        bot.delete_message(message.chat.id, message.message_id)
        bot.send_message(message.chat.id, "Лариса - " + spisok[3] + " в " + tconv(message.date))
        if spisok[3] == "пришла":
            timep = tconv(message.date)
            spisokp[3] = int(timep[0:2])
            spisokpm[3] = int(timep[3:5])
            list["C5"] = timep
            spisok[3] = "ушла"
            spisokr[3] = "пришла"
        elif spisok[3] == "ушла":
            timeu = tconv(message.date)
            spisoku[3] = int(timeu[0:2])
            hp = spisoku[3] - spisokp[3]
            list["G5"] = hp
            spisokum[3] = int(timeu[3:5])
            mp = abs(spisokum[3] - spisokpm[3])
            list["I5"] = mp
            list["E5"] = timeu
            hp1 = int(list1["C5"].value)
            hp1 = hp1 + hp
            list1["C5"] = hp1
            mp1 = int(list1["E5"].value)
            mp1 = mp1 + mp
            list1["E5"] = mp1
            spisok[3] = "пришла"
            spisokr[3] = "ушла"
            if spisokr[0]=="ушёл" and spisokr[1]=="ушла" and spisokr[2]=="ушла" and spisokr[3]=="ушла":
                wb.save('Таблица.xlsx')
                f = open("Таблица.xlsx", "rb")
                bot.send_document(message.chat.id, f)
                f.close()
                os.remove("Таблица.xlsx")
                creatingfile()
                if datetime.datetime.today().weekday()==4:
                    wb1.save('Неделя.xlsx')
                    f = open("Неделя.xlsx", "rb")
                    bot.send_document(message.chat.id, f)
                    f.close()
                    os.remove("Неделя.xlsx")
                    creatingfile()


bot.polling(none_stop = True)